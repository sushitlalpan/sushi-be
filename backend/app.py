# app.py
# Sistema de Login Dual con Flask y Excel (Usuarios, Administradores, Sucursales, Cierre de Caja y Checkador de Asistencia)
from flask import Flask, request, render_template, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import os
import logging
from datetime import datetime, timedelta
import io
import json
from openpyxl import load_workbook

app = Flask(__name__)
# Solución Definitiva para CORS: Configuración explícita para aceptar todas las conexiones.
CORS(app, resources={r"/*": {"origins": "*"}})

# Configuración de logging para monitorear la aplicación
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

basedir = os.path.abspath(os.path.dirname(__file__))
EXCEL_FILE = os.path.join(basedir, 'datos.xlsx')

USERS_SHEET = 'Usuarios'
ADMINS_SHEET = 'Administradores'
BRANCHES_SHEET = 'Sucursales'
CIERRE_CAJA_SHEET = 'Ventas'
EGRESOS_SHEET = 'Egresos'
NOMINA_SHEET = 'Nomina'
# --- NUEVAS HOJAS PARA EL CHECKADOR ---
REGISTRO_SHEET = 'Registro'
PERMISOS_SHEET = 'Permisos'
CONFIG_SHEET = 'Config' # Hoja para guardar la configuración del checkador

# --- Definición de Columnas para cada Hoja ---
# Esto asegura que si una hoja no existe, se cree con la estructura correcta.

CIERRE_CAJA_COLUMNS = [
    'OPERADOR', 'FECHA DE CORTE', 'NO. CIERRE', 'VENTAS TOTALES', 'TARJETA SEGÚN ITPV',
    'DEVOLUCIÓN EN TARJETA', 'TARJETA/ KIWI', 'TRANSFER', 'TOTAL TARJETA REAL', 'EFECTIVO',
    'DEVOLUCIÓN EN EFECTIVO', 'TOTAL EN EFECTIVO REAL', 'DIFERENCIA REVISAR', 'SUCURSAL',
    'NO. PAGOS', 'TICKET PROMEDIO', 'TOTAL COMISIÓN KIWI', 'TOTAL SIN COMISIÓN',
    'TOTAL DE INGRESOS', 'NOTAS', 'FECHA_REGISTRO'
]

EGRESOS_COLUMNS = [
    'FECHA', 'COMPRA_SERVICIO', 'LUGAR_COMPRA', 'CONCEPTO', 'SUCURSAL', 'OPERADOR',
    'CANTIDAD', 'UDM', 'PRECIO', 'FECHA_REGISTRO'
]

NOMINA_COLUMNS = [
    'FECHA', 'OPERADOR', 'SUCURSAL', 'DIAS_LABORADOS', 'CANTIDAD', 'TIPO_NOMINA',
    'NOTAS', 'FECHA_REGISTRO'
]

# --- NUEVAS COLUMNAS PARA EL CHECKADOR ---
# Columnas para la hoja de Usuarios (extendida para el checkador)
USERS_COLUMNS = [
    'Usuario', 'Contraseña', 'Sucursal', 'IDHuella', 'Phone', 'InternalID',
    'WorkArea', 'CustomEntryTime', 'CustomExitTime'
]

# Columnas para la hoja de Registro de Asistencia
REGISTRO_COLUMNS = [
    'Usuario', 'Dia', 'HoraEntrada', 'HoraSalida', 'HorasTrabajadas'
]

# Columnas para la hoja de Permisos y Vacaciones
PERMISOS_COLUMNS = [
    'IDPermiso', 'Empleado', 'Tipo', 'FechaInicio', 'FechaFinal', 'Descripcion'
]

# Columnas para la hoja de Configuración
CONFIG_COLUMNS = [
    'key', 'value'
]

# Reemplaza la función original con esta versión corregida



def load_data_from_sheet(sheet_name, columns):
    try:
        if not os.path.exists(EXCEL_FILE):
            logger.warning(f"Archivo '{EXCEL_FILE}' no encontrado. Se creará uno nuevo.")
            return pd.DataFrame(columns=columns)

        # Cargar libro con openpyxl para obtener valores evaluados de fórmulas
        wb = load_workbook(EXCEL_FILE, data_only=True)
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Hoja '{sheet_name}' no encontrada. Se usará un DataFrame vacío.")
            return pd.DataFrame(columns=columns)

        ws = wb[sheet_name]
        rows = ws.values
        headers = next(rows)
        data = list(rows)

        df = pd.DataFrame(data, columns=headers)
        # Asegurar todas las columnas esperadas
        for col in columns:
            if col not in df.columns:
                df[col] = ''
        df = df[columns].astype(str).fillna('')
        logger.info(f"Hoja '{sheet_name}' cargada con {len(df)} registros.")
        return df

    except Exception as e:
        logger.warning(f"Error cargando hoja '{sheet_name}': {e}")
        return pd.DataFrame(columns=columns)


# Carga todos los datos en un solo diccionario para fácil acceso
data_frames = {
    'users': load_data_from_sheet(USERS_SHEET, USERS_COLUMNS),
    'admins': load_data_from_sheet(ADMINS_SHEET, ['Usuario', 'Contraseña']),
    'branches': load_data_from_sheet(BRANCHES_SHEET, ['Sucursal']),
    'registro_data': load_data_from_sheet(REGISTRO_SHEET, REGISTRO_COLUMNS)
    # Se omiten otras hojas para brevedad, pero en tu código completo deberían estar todas
}
user_data = data_frames['users']
admin_data = data_frames['admins']
branches_data = data_frames['branches']
registro_data = data_frames['registro_data']
        
def load_all_data():
    """Carga todas las hojas de datos en memoria."""
    users = load_data_from_sheet(USERS_SHEET, USERS_COLUMNS)
    admins = load_data_from_sheet(ADMINS_SHEET, ['Usuario', 'Contraseña'])
    branches = load_data_from_sheet(BRANCHES_SHEET, ['Sucursal'])
    ventas = load_data_from_sheet(CIERRE_CAJA_SHEET, CIERRE_CAJA_COLUMNS)
    egresos = load_data_from_sheet(EGRESOS_SHEET, EGRESOS_COLUMNS)
    nomina = load_data_from_sheet(NOMINA_SHEET, NOMINA_COLUMNS)
    # --- Carga de nuevos datos del checkador ---
    registros = load_data_from_sheet(REGISTRO_SHEET, REGISTRO_COLUMNS)
    permisos = load_data_from_sheet(PERMISOS_SHEET, PERMISOS_COLUMNS)
    config = load_data_from_sheet(CONFIG_SHEET, CONFIG_COLUMNS)
    
    return users, admins, branches, ventas, egresos, nomina, registros, permisos, config
    
# ESTA ES LA VERSIÓN CORRECTA QUE DEBE QUEDAR
def save_to_excel(sheet_name, data_frame):
    """
    Guarda un DataFrame específico en su hoja correspondiente sin afectar a las demás.
    """
    try:
        # Usamos el modo 'a' (append/agregar) para editar el archivo sin borrarlo.
        # if_sheet_exists='replace' asegura que solo la hoja que cambiamos se actualice.
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            data_frame.to_excel(writer, sheet_name=sheet_name, index=False)
        logger.info(f"Hoja '{sheet_name}' guardada exitosamente en '{EXCEL_FILE}'.")
        return True
    except FileNotFoundError:
        # Si el archivo no existe, lo crea desde cero.
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='w') as writer:
            data_frame.to_excel(writer, sheet_name=sheet_name, index=False)
        logger.info(f"Archivo '{EXCEL_FILE}' no encontrado. Se ha creado uno nuevo con la hoja '{sheet_name}'.")
        return True
    except PermissionError:
        logger.error(f"Error de permisos al guardar '{EXCEL_FILE}'. Asegúrate de que no esté abierto en otra aplicación.")
        return False
    except Exception as e:
        logger.error(f"Error general al guardar en Excel: {e}")
        return False
        
@app.route('/')
def show_form():
    return jsonify({'status': 'ok', 'message': 'Servidor Sushi Tlalpan funcionando correctamente'})

# Carga inicial de todos los datos al arrancar el servidor
user_data, admin_data, branches_data, cierre_caja_data, egresos_data, nomina_data, registro_data, permisos_data, config_data = load_all_data() 
# SE ELIMINA EL ENDPOINT OPTIONS ESPECÍFICO YA QUE CORS(app) DEBE MANEJARLO
# @app.route('/', methods=['OPTIONS'])
# @app.route('/<path:path>', methods=['OPTIONS'])
# def handle_options(path=None):
#     response = jsonify({'status': 'ok'})
#     response.headers['Access-Control-Allow-Origin'] = '*'
#     response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
#     response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, glycolytic-semipatriotically-alfonso.ngrok-free.app-skip-browser-warning'
#     return response

@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization,true')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response
    
@app.route('/login', methods=['POST'])
def handle_user_login():
    return process_login(user_data, 'usuario')

@app.route('/admin-login', methods=['POST'])
def handle_admin_login():
    return process_login(admin_data, 'administrador')


def process_login(data_source, user_type):
    try:
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if not username or not password:
            return jsonify({'success': False, 'message': 'Usuario y contraseña son requeridos.'})
        
        user_record = data_source[data_source['Usuario'] == username]
        if user_record.empty:
            return jsonify({'success': False, 'message': f'Usuario {user_type} no encontrado.'})
        
        actual_password = str(user_record.iloc[0]['Contraseña']).strip()
        if actual_password == password:
            user_branch = str(user_record.iloc[0].get('Sucursal', '')).strip() if 'Sucursal' in user_record.columns else None
            return jsonify({'success': True, 'message': f'¡Bienvenido {username}!', 'user_type': user_type, 'username': username, 'user_branch': user_branch})
        else:
            return jsonify({'success': False, 'message': 'Contraseña incorrecta.'})
    except Exception as e:
        logger.error(f"Error en login {user_type}: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {str(e)}'})



@app.route('/nomina', methods=['POST'])
def handle_nomina():
    """Procesa el formulario de nómina y guarda los datos en la hoja 'Nomina'."""
    global nomina_data
    try:
        # --- INICIO DE LA CORRECCIÓN 1: Manejo de valores numéricos ---
        # Se obtiene el valor del formulario, se le quitan espacios y si está vacío, se usa '0'.
        dias_laborados_str = request.form.get('dias_laborados', '0').strip()
        cantidad_str = request.form.get('cantidad', '0').strip()

        # Recopilar datos del formulario con claves en MAYÚSCULAS
        form_data = {
            'FECHA': request.form.get('fecha', '').strip(),
            'OPERADOR': request.form.get('operador', '').strip(),
            'SUCURSAL': request.form.get('sucursal', '').strip(),
            'DIAS_LABORADOS': int(dias_laborados_str) if dias_laborados_str else 0,
            'CANTIDAD': float(cantidad_str) if cantidad_str else 0.0,
            'TIPO_NOMINA': request.form.get('tipo_nomina', '').strip(),
            'NOTAS': request.form.get('notas', '').strip(),
            'FECHA_REGISTRO': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        # --- FIN DE LA CORRECIÓN 1 ---

        # Validación básica de campos requeridos
        required_fields = ['FECHA', 'OPERADOR', 'SUCURSAL', 'TIPO_NOMINA']
        if not all(form_data[field] for field in required_fields) or form_data['DIAS_LABORADOS'] <= 0 or form_data['CANTIDAD'] <= 0:
            return jsonify({'success': False, 'message': 'Todos los campos son requeridos y los valores numéricos deben ser mayores a cero.'})

        nuevo_registro = pd.DataFrame([form_data])
        nomina_data = pd.concat([nomina_data, nuevo_registro], ignore_index=True)
        
        # --- CORRECCIÓN 2: Llamada correcta a save_to_excel ---
        if save_to_excel(NOMINA_SHEET, nomina_data):
            return jsonify({'success': True, 'message': 'Registro de nómina guardado exitosamente.'})
        else:
            nomina_data = nomina_data.iloc[:-1] # Revertir si falla el guardado
            return jsonify({'success': False, 'message': 'Error crítico al guardar en Excel.'})

    except (ValueError, TypeError) as e:
        logger.error(f"Error de conversión de datos en el formulario de nómina: {e}")
        return jsonify({'success': False, 'message': 'Error en el formato de datos numéricos (Dias/Cantidad).'})
    except Exception as e:
        logger.error(f"Error al procesar la nómina: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {e}'})


@app.route('/nomina/list', methods=['GET'])
def list_nomina():
    """Lista todos los registros de nómina para el panel de gráficas."""
    if nomina_data is None or nomina_data.empty:
        return jsonify({'registros': [], 'total': 0})

    # Crear una copia para manipularla de forma segura
    data_copy = nomina_data.copy()

    # Reemplazar valores NaN de pandas por None para que el JSON sea válido
    data_copy = data_copy.where(pd.notnull(data_copy), None)

    # Convertir el DataFrame a una lista de diccionarios (formato JSON)
    registros = data_copy.to_dict('records')
    
    return jsonify({'registros': registros, 'total': len(registros)})

@app.route('/datos/generales', methods=['GET'])
def get_datos_generales():
    """
    Endpoint unificado para devolver los datos de Ventas, Egresos y Nómina.
    """
    # Preparar los datos de cada hoja, asegurándose de que no sean None
    ventas = cierre_caja_data.where(pd.notnull(cierre_caja_data), None).to_dict('records') if cierre_caja_data is not None else []
    egresos = egresos_data.where(pd.notnull(egresos_data), None).to_dict('records') if egresos_data is not None else []
    nomina = nomina_data.where(pd.notnull(nomina_data), None).to_dict('records') if nomina_data is not None else []

    # Devolver todos los datos en un solo objeto JSON
    return jsonify({
        'ventas': ventas,
        'egresos': egresos,
        'nomina': nomina
    })    

# ESTA ES LA VERSIÓN CORRECTA QUE DEBES CONSERVAR
def process_login(data_source, user_type):
    try:
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        if not username or not password:
            return jsonify({'success': False, 'message': 'Usuario y contraseña son requeridos.'})

        user_record = data_source[data_source['Usuario'] == username]

        if user_record.empty:
            return jsonify({'success': False, 'message': f'Usuario {user_type} no encontrado.'})

        actual_password = str(user_record.iloc[0]['Contraseña']).strip()

        if actual_password == password:
            user_branch = str(user_record.iloc[0].get('Sucursal', '')).strip() if 'Sucursal' in user_record.columns else None
            return jsonify({
                'success': True,
                'message': f'¡Bienvenido {username}!',
                'user_type': user_type,
                'username': username,
                'user_branch': user_branch
            })
        else:
            return jsonify({'success': False, 'message': 'Contraseña incorrecta.'})

    except Exception as e:
        logger.error(f"Error en login {user_type}: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {str(e)}'})


@app.route('/users', methods=['GET'])
def list_users():
    """
    Endpoint para listar usuarios con su contraseña y sucursal, 
    en el formato que espera el frontend de gestión de usuarios.
    """
    if user_data is None or user_data.empty:
        return jsonify({'users': [], 'total': 0})
    
    # Crea una copia para evitar modificar los datos originales en memoria
    data_copy = user_data.copy()

    # Asegura que las columnas existan para evitar errores
    for col in ['Usuario', 'Contraseña', 'Sucursal']:
        if col not in data_copy.columns:
            data_copy[col] = 'No disponible'

    # Convierte el DataFrame a una lista de diccionarios
    # [{"Usuario": "...", "Contraseña": "...", "Sucursal": "..."}]
    records = data_copy.to_dict('records')

    # Transforma las claves a minúsculas para que coincidan con el JavaScript
    # y crea la lista final: [{"usuario": "...", "contrasena": "...", "sucursal": "..."}]
    users_list = []
    for record in records:
        users_list.append({
            'usuario': record.get('Usuario'),
            'contrasena': record.get('Contraseña'),
            'sucursal': record.get('Sucursal')
        })

    return jsonify({'users': users_list, 'total': len(users_list)})

@app.route('/cierre-caja', methods=['POST'])
def handle_cierre_caja():
    """Procesa el formulario de ventas y guarda los datos."""
    global cierre_caja_data
    try:
        fecha_str = request.form.get('fecha', '').strip()
        fecha_final = None
        if fecha_str:
            fecha_final = datetime.strptime(fecha_str, '%Y-%m-%d')

        # --- INICIO DE LA CORRECCIÓN ---
        # 1. Obtenemos los valores de 'TARJETA/ KIWI' y 'TRANSFER' como números.
        tarjeta_kiwi_valor = float(request.form.get('tarjetaKiwi', '0'))
        transfer_valor = float(request.form.get('transferencias', '0'))

        # 2. Calculamos el 'TOTAL TARJETA REAL' en el servidor.
        total_tarjeta_real_calculado = tarjeta_kiwi_valor + transfer_valor
        # --- FIN DE LA CORRECCIÓN ---

        form_data = {
            'OPERADOR': request.form.get('operador', '').strip(),
            'FECHA DE CORTE': fecha_final,
            'NO. CIERRE': request.form.get('numeroCierre', '').strip(),
            'VENTAS TOTALES': float(request.form.get('ventasTotales', '0')),
            'TARJETA SEGÚN ITPV': float(request.form.get('tarjetaITPV', '0')),
            'DEVOLUCIÓN EN TARJETA': float(request.form.get('devolucionTarjeta', '0')),
            'TARJETA/ KIWI': tarjeta_kiwi_valor, # Guardamos el valor numérico
            'TRANSFER': transfer_valor,         # Guardamos el valor numérico
            'TOTAL TARJETA REAL': total_tarjeta_real_calculado, # <-- CORREGIDO: Usamos el valor calculado
            'EFECTIVO': float(request.form.get('efectivo', '0')),
            'DEVOLUCIÓN EN EFECTIVO': float(request.form.get('devolucionEfectivo', '0')),
            'TOTAL EN EFECTIVO REAL': float(request.form.get('totalEfectivoReal', '0')),
            'DIFERENCIA REVISAR': request.form.get('diferenciaRevisar', '0'),
            'SUCURSAL': request.form.get('sucursal', '').strip(),
            'NO. PAGOS': int(request.form.get('numeroPagos', '0') or 0),
            'TICKET PROMEDIO': float(request.form.get('ticketPromedio', '0')),
            'TOTAL COMISIÓN KIWI': float(request.form.get('totalComisionKiwi', '0')),
            'TOTAL SIN COMISIÓN': float(request.form.get('totalSinComision', '0')),
            'TOTAL DE INGRESOS': float(request.form.get('totalIngresos', '0')),
            'NOTAS': request.form.get('notas', ''),
            'FECHA_REGISTRO': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        if not all([form_data['OPERADOR'], form_data['FECHA DE CORTE'], form_data['NO. CIERRE'], form_data['SUCURSAL']]):
            return jsonify({'success': False, 'message': 'Operador, fecha, N° de cierre y sucursal son requeridos.'})

        nuevo_registro = pd.DataFrame([form_data])
        
        # Este bloque ya no es necesario si guardamos los valores como números
        # for col in ['TARJETA/ KIWI', 'TRANSFER']:
        #     if col in nuevo_registro.columns:
        #         nuevo_registro[col] = nuevo_registro[col].astype(object)

        cierre_caja_data = pd.concat([cierre_caja_data, nuevo_registro], ignore_index=True)
        
        if save_to_excel(CIERRE_CAJA_SHEET, cierre_caja_data):
            return jsonify({'success': True, 'message': 'Venta guardada exitosamente.'})
        else:
            cierre_caja_data = cierre_caja_data.iloc[:-1] 
            return jsonify({'success': False, 'message': 'Error crítico al guardar en Excel.'})

    except ValueError as e:
        logger.error(f"Error de conversión de datos en el formulario: {e}")
        return jsonify({'success': False, 'message': 'Error en el formato de datos numéricos.'})
    except Exception as e:
        logger.error(f"Error al procesar la venta: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {e}'})

@app.route('/cierre-caja/download', methods=['POST'])
def download_cierre_caja():
    """
    Genera y envía un archivo Excel con formato condicional (colores y observaciones)
    basado en los datos enviados desde el cliente.
    """
    try:
        # Recibe los datos de la tabla (incluyendo estado y observaciones) desde el frontend
        rows_data = request.get_json()

        if not rows_data:
            return jsonify({'success': False, 'message': 'No se recibieron datos para exportar.'}), 400

        # Convierte los datos JSON en un DataFrame de pandas
        df = pd.DataFrame(rows_data)
        
        # Opcional: Reordena o elimina columnas si es necesario para el reporte
        # Por ejemplo, podemos quitar la columna 'status' que solo se usa para formato
        if 'status' in df.columns:
            # Mapea el estado a una columna más descriptiva si lo deseas
            df['ESTADO'] = df['status'].apply(lambda s: 'Revisado' if s == 'green' else ('Alerta' if s == 'red' else 'Normal'))
            # Elimina la columna original de status y el id interno
            df = df.drop(columns=['status', 'id'])


        # Crea un buffer de bytes en memoria para el archivo Excel
        output = io.BytesIO()

        # Usa ExcelWriter para poder aplicar formatos
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Ventas')

            # Obtiene el workbook y la worksheet para trabajar con ellos
            workbook = writer.book
            worksheet = writer.sheets['Ventas']

            # Define los formatos de color para las filas
            red_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
            green_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})

            # Itera sobre los datos originales para aplicar el formato fila por fila
            # Se suma 1 a 'idx' porque las filas de Excel son 1-indexadas y la fila 0 es la cabecera
            for idx, row_data in enumerate(rows_data):
                if row_data.get('status') == 'red':
                    worksheet.set_row(idx + 1, None, red_format)
                elif row_data.get('status') == 'green':
                    worksheet.set_row(idx + 1, None, green_format)
            
            # Ajusta el ancho de las columnas para mejor legibilidad
            for i, col in enumerate(df.columns):
                width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(i, i, width)

        # Regresa al inicio del buffer para poder leer su contenido
        output.seek(0)

        # Envía el archivo al cliente como una descarga adjunta
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='reporte_ventas_formateado.xlsx'
        )

    except Exception as e:
        logger.error(f"Error al generar el archivo Excel de ventas: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {e}'}), 500


@app.route('/cierre-caja/list', methods=['GET'])
def list_cierre_caja():
    """Lista todos los registros de ventas, ideal para el panel de gráficas."""
    if cierre_caja_data is None or cierre_caja_data.empty:
        return jsonify({'registros': [], 'total': 0})

    # Prepara una copia para no modificar el DataFrame global
    data_copy = cierre_caja_data.copy()

    # --- INICIO DE LA CORRECCIÓN ---
    # Este bloque es crucial. Lee cualquier formato de fecha del Excel 
    # (incluyendo '2025-07-23 00:00:00') y lo convierte a un texto simple '2025-07-23'.
    if 'FECHA DE CORTE' in data_copy.columns:
         # Convierte la columna a objetos de fecha de pandas.
         # 'errors=coerce' maneja celdas vacías o con errores.
         dates = pd.to_datetime(data_copy['FECHA DE CORTE'], errors='coerce')
         
         # Formatea las fechas válidas al string 'YYYY-MM-DD', eliminando la hora.
         data_copy['FECHA DE CORTE'] = dates.dt.strftime('%Y-%m-%d')
    # --- FIN DE LA CORRECCIÓN ---

    # Reemplaza cualquier valor nulo o no válido (NaT) por None,
    # que se convierte en 'null' en el JSON.
    data_copy = data_copy.where(pd.notnull(data_copy), None)

    registros = data_copy.to_dict('records')
    return jsonify({'registros': registros, 'total': len(registros)})

@app.route('/cierre-caja/delete', methods=['POST'])
def delete_cierre_caja():
    """
    Elimina registros de cierre de caja basados en una lista de IDs (FECHA_REGISTRO).
    """
    global cierre_caja_data
    
    try:
        # Obtener la lista de identificadores únicos del cuerpo de la solicitud
        data = request.get_json()
        ids_to_delete = data.get('ids', [])

        if not ids_to_delete:
            return jsonify({'success': False, 'message': 'No se proporcionaron registros para eliminar.'}), 400

        # Contar cuántos registros existen antes de borrar
        initial_count = len(cierre_caja_data)
        
        # Filtrar el DataFrame, conservando solo las filas cuyo 'FECHA_REGISTRO' NO está en la lista de IDs a eliminar.
        # Esto asegura que se eliminen los registros correctos de forma segura.
        cierre_caja_data = cierre_caja_data[~cierre_caja_data['FECHA_REGISTRO'].isin(ids_to_delete)]
        
        # Contar cuántos registros se eliminaron
        deleted_count = initial_count - len(cierre_caja_data)

        if deleted_count == 0:
            return jsonify({'success': False, 'message': 'No se encontraron los registros especificados para eliminar.'}), 404

        # Guardar los cambios en el archivo Excel
        if save_to_excel(CIERRE_CAJA_SHEET, cierre_caja_data):
            logger.info(f"Se eliminaron {deleted_count} registros de cierre de caja.")
            return jsonify({'success': True, 'message': f'{deleted_count} registros eliminados exitosamente.'})
        else:
            # Si el guardado falla, es crucial recargar los datos para evitar inconsistencias.
            cierre_caja_data, _, _, _, _, _, _, _, _ = load_all_data() # Se recargan todos los datos
            return jsonify({'success': False, 'message': 'Error crítico al guardar los cambios en Excel. Los datos han sido revertidos.'}), 500

    except Exception as e:
        logger.error(f"Error en /cierre-caja/delete: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {e}'}), 500

@app.route('/cierre-caja/search', methods=['POST'])
def search_cierre_caja():
    """
    Busca registros de cierre de caja por criterios
    """
    if cierre_caja_data is None or cierre_caja_data.empty:
        return jsonify({
            'error': 'No hay registros de cierre de caja',
            'registros': [],
            'total': 0
        })
    
    try:
        # Obtener criterios de búsqueda
        operador = request.form.get('operador', '').strip()
        fecha_inicio = request.form.get('fecha_inicio', '').strip()
        fecha_fin = request.form.get('fecha_fin', '').strip()
        sucursal = request.form.get('sucursal', '').strip()
        
        # Filtrar datos
        filtered_data = cierre_caja_data.copy()
        
        if operador:
            filtered_data = filtered_data[filtered_data['OPERADOR'] == operador]
        
        if sucursal:
            filtered_data = filtered_data[filtered_data['SUCURSAL'] == sucursal]
        
        if fecha_inicio and fecha_fin:
            filtered_data = filtered_data[
                (filtered_data['FECHA DE CORTE'] >= fecha_inicio) & 
                (filtered_data['FECHA DE CORTE'] <= fecha_fin)
            ]
        
        # Convertir a lista de diccionarios
        registros = filtered_data.to_dict('records')
        
        return jsonify({
            'registros': registros,
            'total': len(registros),
            'criterios': {
                'operador': operador,
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'sucursal': sucursal
            }
        })
        
    except Exception as e:
        logger.error(f"Error al buscar cierre de caja: {e}")
        return jsonify({
            'success': False,
            'message': f'Error al buscar: {str(e)}'
        })


@app.route('/nomina/delete', methods=['POST'])
def delete_nomina():
    """
    Elimina registros de la hoja de nómina basados en una lista de IDs (FECHA_REGISTRO).
    """
    global nomina_data
    
    try:
        # Obtener la lista de identificadores únicos del cuerpo de la solicitud
        data = request.get_json()
        ids_to_delete = data.get('ids', [])

        if not ids_to_delete:
            return jsonify({'success': False, 'message': 'No se proporcionaron registros para eliminar.'}), 400

        # Contar cuántos registros existen antes de borrar
        initial_count = len(nomina_data)
        
        # Filtrar el DataFrame, conservando solo las filas cuyo 'FECHA_REGISTRO' NO está en la lista de IDs a eliminar.
        nomina_data = nomina_data[~nomina_data['FECHA_REGISTRO'].isin(ids_to_delete)]
        
        # Contar cuántos registros se eliminaron
        deleted_count = initial_count - len(nomina_data)

        if deleted_count == 0:
            return jsonify({'success': False, 'message': 'No se encontraron los registros especificados para eliminar.'}), 404

        # Guardar los cambios en el archivo Excel
        if save_to_excel(NOMINA_SHEET, nomina_data):
            logger.info(f"Se eliminaron {deleted_count} registros de nómina.")
            return jsonify({'success': True, 'message': f'{deleted_count} registros de nómina eliminados exitosamente.'})
        else:
            # Si el guardado falla, es crucial recargar los datos para evitar inconsistencias.
            _, _, _, _, _, nomina_data, _, _, _ = load_all_data()
            return jsonify({'success': False, 'message': 'Error crítico al guardar los cambios en Excel. Los datos han sido revertidos.'}), 500

    except Exception as e:
        logger.error(f"Error en /nomina/delete: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {e}'}), 500


  
@app.route('/egresos/delete', methods=['POST'])
def delete_egresos():
    """
    Elimina registros de la hoja de Egresos basados en una lista de IDs (FECHA_REGISTRO).
    """
    global egresos_data
    
    try:
        # Obtener la lista de identificadores únicos del cuerpo de la solicitud
        data = request.get_json()
        ids_to_delete = data.get('ids', [])

        if not ids_to_delete:
            return jsonify({'success': False, 'message': 'No se proporcionaron registros para eliminar.'}), 400

        # Contar cuántos registros existen antes de borrar
        initial_count = len(egresos_data)
        
        # Filtrar el DataFrame, conservando solo las filas cuyo 'FECHA_REGISTRO' NO está en la lista de IDs a eliminar.
        egresos_data = egresos_data[~egresos_data['FECHA_REGISTRO'].isin(ids_to_delete)]
        
        # Contar cuántos registros se eliminaron
        deleted_count = initial_count - len(egresos_data)

        if deleted_count == 0:
            return jsonify({'success': False, 'message': 'No se encontraron los registros especificados para eliminar.'}), 404

        # Guardar los cambios en el archivo Excel
        if save_to_excel(EGRESOS_SHEET, egresos_data):
            logger.info(f"Se eliminaron {deleted_count} registros de egresos.")
            return jsonify({'success': True, 'message': f'{deleted_count} registros de egresos eliminados exitosamente.'})
        else:
            # Si el guardado falla, se recargan los datos para evitar inconsistencias.
            _, _, _, _, egresos_data, _, _, _, _ = load_all_data()
            return jsonify({'success': False, 'message': 'Error crítico al guardar los cambios en Excel. Los datos han sido revertidos.'}), 500

    except Exception as e:
        logger.error(f"Error en /egresos/delete: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {e}'}), 500


      

@app.route('/registro-entrada', methods=['POST'])
def registro_entrada():
    """Añade una nueva fila de entrada para un usuario."""
    global registro_data
    try:
        usuario = request.form.get('Usuario')
        dia = request.form.get('Dia')
        hora_entrada = request.form.get('HoraEntrada')

        if not all([usuario, dia, hora_entrada]):
            return jsonify({'success': False, 'message': 'Faltan datos para el registro de entrada.'}), 400

        user_today_records = registro_data[(registro_data['Usuario'] == usuario) & (registro_data['Dia'] == dia)]
        open_shifts = user_today_records[pd.isna(user_today_records['HoraSalida']) | (user_today_records['HoraSalida'] == '')]
        
        if not open_shifts.empty:
            return jsonify({'success': False, 'message': 'Ya tienes un turno de entrada registrado. Primero debes registrar tu salida.'}), 409

        nuevo_registro_dict = {
            'Usuario': usuario, 'Dia': dia, 'HoraEntrada': hora_entrada,
            'HoraSalida': '', 'HorasTrabajadas': ''
        }
        
        nuevo_registro_df = pd.DataFrame([nuevo_registro_dict])
        registro_data = pd.concat([registro_data, nuevo_registro_df], ignore_index=True)
        
        # --- MODIFICACIÓN CLAVE ---
        if save_to_excel(REGISTRO_SHEET, registro_data):
            current_turn = len(registro_data[(registro_data['Usuario'] == usuario) & (registro_data['Dia'] == dia)])
            return jsonify({
                'success': True, 
                'message': 'Entrada registrada con éxito.',
                'newState': 'awaiting_exit',
                'turn': current_turn
            })
        else:
            registro_data = registro_data.iloc[:-1]
            return jsonify({'success': False, 'message': 'Error al guardar el registro.'}), 500

    except Exception as e:
        logger.error(f"Error en /registro-entrada: {e}")
        return jsonify({'success': False, 'message': f'Error interno: {e}'}), 500


@app.route('/registro-salida', methods=['POST'])
def registro_salida():
    """Registra la hora de salida en el último turno abierto del día."""
    global registro_data
    try:
        usuario = request.form.get('Usuario')
        dia = request.form.get('Dia')
        hora_salida = request.form.get('HoraSalida')

        if not all([usuario, dia, hora_salida]):
            return jsonify({'success': False, 'message': 'Faltan datos para el registro de salida.'}), 400

        # Busca los registros del usuario para el día de hoy, ignorando mayúsculas/minúsculas
        user_today_records = registro_data[(registro_data['Usuario'].str.upper() == usuario.upper()) & (registro_data['Dia'] == dia)]
        
        # Encuentra los turnos que no tienen hora de salida
        open_shifts = user_today_records[user_today_records['HoraSalida'].astype(str).str.strip() == '']

        if open_shifts.empty:
            return jsonify({'success': False, 'message': 'No se encontró un turno abierto para registrar la salida.'}), 404

        # Obtiene el índice del último turno abierto para actualizarlo
        idx_to_update = open_shifts.index[-1]
        hora_entrada_str = registro_data.loc[idx_to_update, 'HoraEntrada']
        
        horas_trabajadas = "00:00:00" # Valor por defecto
        try:
            FMT = '%H:%M:%S'
            # Calcula la diferencia de tiempo
            tdelta = datetime.strptime(hora_salida, FMT) - datetime.strptime(hora_entrada_str, FMT)
            
            # Si la salida es al día siguiente, ajusta el cálculo
            if tdelta.total_seconds() < 0: 
                tdelta += timedelta(days=1)
            
            # --- INICIO DE LA CORRECCIÓN ---
            # Descompone el total de segundos en horas, minutos y segundos
            total_seconds = int(tdelta.total_seconds())
            hours, remainder = divmod(total_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            
            # Formatea el resultado en formato HH:MM:SS
            horas_trabajadas = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
            # --- FIN DE LA CORRECCIÓN ---

        except (ValueError, TypeError) as e:
            logger.error(f"Error calculando horas trabajadas para {usuario}: {e}")
            horas_trabajadas = "Error de cálculo"

        # Actualiza el DataFrame en memoria con la hora de salida y las horas calculadas
        registro_data.loc[idx_to_update, 'HoraSalida'] = hora_salida
        registro_data.loc[idx_to_update, 'HorasTrabajadas'] = horas_trabajadas
        
        # Guarda el DataFrame actualizado en el archivo Excel
        if save_to_excel(REGISTRO_SHEET, registro_data):
            next_turn = len(user_today_records) + 1
            return jsonify({
                'success': True, 
                'message': 'Salida registrada. ¡Hasta luego!',
                'newState': 'awaiting_entry',
                'turn': next_turn
            })
        else:
            # Si falla el guardado, se podría revertir el cambio en memoria (opcional)
            # load_all_data() 
            return jsonify({'success': False, 'message': 'Error al guardar el registro de salida.'}), 500
            
    except Exception as e:
        logger.error(f"Error en /registro-salida: {e}")
        return jsonify({'success': False, 'message': f'Error interno: {e}'}), 500


@app.route('/check-status')
def check_status():
    usuario = request.args.get('usuario')
    if not usuario:
        return jsonify({"error": "Falta el parámetro 'usuario'"}), 400
    try:
        # --- CORRECCIÓN CLAVE: Recargar los datos frescos desde Excel ---
        # Esto asegura que siempre se consulta el estado más reciente guardado en el archivo.
        df = load_data_from_sheet(REGISTRO_SHEET, REGISTRO_COLUMNS)

        today_str = datetime.now().strftime('%Y-%m-%d')
        # Se usa la variable local 'df' que contiene los datos actualizados
        user_today_records = df[(df['Usuario'].str.upper() == usuario.upper()) & (df['Dia'] == today_str)]
        
        total_records_today = len(user_today_records)
        
        if user_today_records.empty:
            return jsonify({"status": "awaiting_entry", "turn": 1})
        
        last_record = user_today_records.iloc[-1]
        
        # La lógica para determinar el estado ya era correcta y no cambia
        if not str(last_record['HoraSalida']).strip():
            return jsonify({"status": "awaiting_exit", "turn": total_records_today})
        else:
            return jsonify({"status": "awaiting_entry", "turn": total_records_today + 1})
            
    except Exception as e:
        logger.error(f"Error en /check-status para el usuario {usuario}: {e}")
        return jsonify({'success': False, 'message': f'Error interno: {e}'}), 500
        
        
@app.route('/users-with-passwords', methods=['GET'])
def list_users_with_passwords():
    """
    Endpoint para listar usuarios con sus contraseñas (solo para gestión)
    """
    if user_data is None or user_data.empty:
        return jsonify({
            'error': 'No se pudieron cargar los usuarios',
            'users': {}
        })
    
    users_dict = {}
    # Se incluye la sucursal para gestión si la columna existe
    if 'Sucursal' in user_data.columns:
        for _, row in user_data.iterrows():
            users_dict[row['Usuario']] = {'password': row['Contraseña'], 'branch': row['Sucursal']}
    else:
        for _, row in user_data.iterrows():
            users_dict[row['Usuario']] = {'password': row['Contraseña']}
    
    return jsonify({
        'users': users_dict,
        'total': len(users_dict),
        'type': 'usuarios_con_contraseñas'
    })

@app.route('/permissions/backup', methods=['GET'])
def backup_permissions():
    """Genera y envía un respaldo de la hoja de permisos en un nuevo archivo Excel."""
    try:
        if permisos_data.empty:
            # Se devuelve un mensaje si no hay datos que respaldar.
            return jsonify({'success': False, 'message': 'No hay datos de permisos para respaldar.'}), 404

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            permisos_data.to_excel(writer, index=False, sheet_name='Respaldo Permisos')
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y-%m-%d')
        filename = f"Respaldo_Permisos_{timestamp}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"Error al generar el respaldo de permisos: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {e}'}), 500
        


@app.route('/add-user', methods=['POST'])
def add_user():
    """
    Endpoint para agregar un nuevo usuario al Excel
    """
    global user_data
    
    try:
        # --- CAMBIO DE MÉTODO ---
        # 1. Obtenemos el cuerpo de la solicitud como un JSON
        data = request.get_json()
        
        # 2. Extraemos los datos del diccionario resultante
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        branch = data.get('sucursal', '').strip() # Usamos 'sucursal' como envía el frontend
        
        # 3. La validación ahora funcionará correctamente
        if not username or not password or not branch:
            return jsonify({
                'success': False,
                'message': 'Usuario, contraseña y sucursal son requeridos.'
            })
        
        if user_data is None:
            return jsonify({
                'success': False,
                'message': 'Error del servidor: No se pudo cargar el archivo Excel.'
            })
        
        if username in user_data['Usuario'].values:
            return jsonify({
                'success': False,
                'message': f'El usuario "{username}" ya existe.'
            })
        
        new_row = pd.DataFrame({'Usuario': [username], 'Contraseña': [password], 'Sucursal': [branch]})
        user_data = pd.concat([user_data, new_row], ignore_index=True)
        
        if save_to_excel(USERS_SHEET, user_data):
            logger.info(f"Usuario agregado exitosamente: {username} en sucursal: {branch}")
            return jsonify({
                'success': True,
                'message': f'Usuario "{username}" agregado exitosamente en {branch}.',
                'username': username,
                'branch': branch
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Error al guardar en el archivo Excel.'
            })
        
    except Exception as e:
        logger.error(f"Error al agregar usuario: {e}")
        return jsonify({
            'success': False,
            'message': f'Error interno del servidor: {str(e)}'
        })


@app.route('/user-branches', methods=['GET'])
def get_user_branches():
    """
    Endpoint para obtener el mapeo usuario-sucursal basado en el Excel
    """
    try:
        # Verificar si los datos de usuarios están disponibles
        if user_data is None or user_data.empty:
            return jsonify({
                'success': False,
                'message': 'No se pudieron cargar los datos de usuarios.',
                'userBranches': {}
            })
        
        # Crear el mapeo usuario-sucursal
        user_branches = {}
        
        # Iterar sobre los datos de usuarios
        for _, row in user_data.iterrows():
            username = str(row['Usuario']).strip()
            # Verificar si existe la columna Sucursal
            if 'Sucursal' in user_data.columns and pd.notna(row['Sucursal']):
                branch = str(row['Sucursal']).strip()
                if branch and branch != 'nan' and branch != '':
                    user_branches[username] = branch
        
        logger.info(f"Mapeo usuario-sucursal generado: {user_branches}")
        
        return jsonify({
            'success': True,
            'userBranches': user_branches,
            'total': len(user_branches)
        })
        
    except Exception as e:
        logger.error(f"Error al generar mapeo usuario-sucursal: {e}")
        return jsonify({
            'success': False,
            'message': f'Error interno del servidor: {str(e)}',
            'userBranches': {}
        })



# app.py

@app.route('/delete-user', methods=['POST'])
def delete_user():
    """
    Endpoint para eliminar un usuario del Excel
    """
    global user_data
    
    try:
        username = request.form.get('username', '').strip()
        
        if not username:
            return jsonify({
                'success': False,
                'message': 'Nombre de usuario es requerido.'
            })
        
        if user_data is None or user_data.empty:
            return jsonify({
                'success': False,
                'message': 'Error del servidor: No se pudo cargar el archivo Excel.'
            })
        
        if username not in user_data['Usuario'].values:
            return jsonify({
                'success': False,
                'message': f'El usuario "{username}" no existe.'
            })
        
        user_data = user_data[user_data['Usuario'] != username]
        
        # --- LÍNEA CORREGIDA ---
        if save_to_excel(USERS_SHEET, user_data):
            logger.info(f"Usuario eliminado exitosamente: {username}")
            return jsonify({
                'success': True,
                'message': f'Usuario "{username}" eliminado exitosamente.',
                'username': username
            })
        else:
            # Si falla, se recargan los datos para mantener la consistencia
            load_all_data() 
            return jsonify({
                'success': False,
                'message': 'Error al guardar en el archivo Excel. Los cambios han sido revertidos.'
            })
        
    except Exception as e:
        logger.error(f"Error al eliminar usuario: {e}")
        return jsonify({
            'success': False,
            'message': f'Error interno del servidor: {str(e)}'
        })

@app.route('/modify-user', methods=['POST'])
def modify_user():
    """
    Endpoint para modificar un usuario existente (nombre, contraseña o sucursal).
    """
    global user_data
    
    try:
        username = request.form.get('username', '').strip()
        new_username = request.form.get('new_username', '').strip()
        new_password = request.form.get('new_password', '').strip()
        new_sucursal = request.form.get('new_sucursal', '').strip()
        
        if not username:
            return jsonify({'success': False, 'message': 'No se especificó el usuario a modificar.'})
        
        if not any([new_username, new_password, new_sucursal]):
            return jsonify({'success': False, 'message': 'No se proporcionaron datos para modificar.'})
            
        user_exists = user_data['Usuario'] == username
        if not user_exists.any():
            return jsonify({'success': False, 'message': f'El usuario "{username}" no fue encontrado.'})
        
        if new_username and new_username != username and new_username in user_data['Usuario'].values:
            return jsonify({'success': False, 'message': f'El nuevo nombre de usuario "{new_username}" ya está en uso.'})

        user_index = user_data.index[user_exists].tolist()[0]
        
        if new_username:
            user_data.loc[user_index, 'Usuario'] = new_username
        if new_password:
            user_data.loc[user_index, 'Contraseña'] = new_password
        if new_sucursal:
            user_data.loc[user_index, 'Sucursal'] = new_sucursal

        # --- LÍNEA CORREGIDA ---
        if save_to_excel(USERS_SHEET, user_data):
            logger.info(f"Usuario '{username}' modificado exitosamente.")
            return jsonify({'success': True, 'message': f'Usuario "{username}" modificado exitosamente.'})
        else:
            load_all_data()
            return jsonify({'success': False, 'message': 'Error crítico al guardar los cambios. Los datos han sido revertidos.'})

    except Exception as e:
        logger.error(f"Error al modificar usuario: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {str(e)}'})

@app.route('/admins', methods=['GET'])
def list_admins():
    """
    Endpoint para listar administradores
    """
    if admin_data is None or admin_data.empty:
        return jsonify({
            'error': 'No se pudieron cargar los administradores',
            'admins': [],
            'total': 0
        })
    
    admins_list = admin_data['Usuario'].tolist()
    return jsonify({
        'admins': admins_list,
        'total': len(admins_list),
        'type': 'administradores'
    })

@app.route('/admins-with-passwords', methods=['GET'])
def list_admins_with_passwords():
    """
    Endpoint para listar administradores con sus contraseñas (solo para gestión)
    """
    if admin_data is None or admin_data.empty:
        return jsonify({
            'error': 'No se pudieron cargar los administradores',
            'admins': {}
        })
    
    admins_dict = {}
    for _, row in admin_data.iterrows():
        admins_dict[row['Usuario']] = row['Contraseña']
    
    return jsonify({
        'admins': admins_dict,
        'total': len(admins_dict),
        'type': 'administradores_con_contraseñas'
    })

# app.py

@app.route('/add-admin', methods=['POST'])
def add_admin():
    """
    Endpoint para agregar un nuevo administrador al Excel
    """
    global admin_data
    
    try:
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        if not username or not password:
            return jsonify({
                'success': False,
                'message': 'Usuario y contraseña son requeridos.'
            })
        
        if admin_data is None:
            return jsonify({
                'success': False,
                'message': 'Error del servidor: No se pudo cargar el archivo Excel.'
            })
        
        if username in admin_data['Usuario'].values:
            return jsonify({
                'success': False,
                'message': f'El administrador "{username}" ya existe.'
            })
        
        new_row = pd.DataFrame({'Usuario': [username], 'Contraseña': [password]})
        admin_data = pd.concat([admin_data, new_row], ignore_index=True)
        
        # --- LÍNEA CORREGIDA ---
        if save_to_excel(ADMINS_SHEET, admin_data):
            logger.info(f"Administrador agregado exitosamente: {username}")
            return jsonify({
                'success': True,
                'message': f'Administrador "{username}" agregado exitosamente.',
                'username': username
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Error al guardar en el archivo Excel. Verifica que el archivo no esté abierto en otra aplicación.'
            })
        
    except Exception as e:
        logger.error(f"Error al agregar administrador: {e}")
        return jsonify({
            'success': False,
            'message': f'Error interno del servidor: {str(e)}'
        })

# app.py

@app.route('/delete-admin', methods=['POST'])
def delete_admin():
    """
    Endpoint para eliminar un administrador del Excel
    """
    global admin_data
    
    try:
        username = request.form.get('username', '').strip()
        
        if not username:
            return jsonify({
                'success': False,
                'message': 'Nombre de administrador es requerido.'
            })
        
        if admin_data is None or admin_data.empty:
            return jsonify({
                'success': False,
                'message': 'Error del servidor: No se pudo cargar el archivo Excel.'
            })
        
        if username not in admin_data['Usuario'].values:
            return jsonify({
                'success': False,
                'message': f'El administrador "{username}" no existe.'
            })
        
        admin_data = admin_data[admin_data['Usuario'] != username]
        
        # --- LÍNEA CORREGIDA ---
        if save_to_excel(ADMINS_SHEET, admin_data):
            logger.info(f"Administrador eliminado exitosamente: {username}")
            return jsonify({
                'success': True,
                'message': f'Administrador "{username}" eliminado exitosamente.',
                'username': username
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Error al guardar en el archivo Excel. Verifica que el archivo no esté abierto en otra aplicación.'
            })
        
    except Exception as e:
        logger.error(f"Error al eliminar administrador: {e}")
        return jsonify({
            'success': False,
            'message': f'Error interno del servidor: {str(e)}'
        })

@app.route('/branches', methods=['GET'])
def list_branches():
    """
    Endpoint para listar sucursales
    """
    if branches_data is None or branches_data.empty:
        return jsonify({
            'error': 'No se pudieron cargar las sucursales',
            'branches': [],
            'total': 0
        })
    
    branches_list = branches_data['Sucursal'].tolist()
    return jsonify({
        'branches': branches_list,
        'total': len(branches_list),
        'type': 'sucursales'
    })

# CORRECTO
@app.route('/add-branch', methods=['POST'])
def add_branch():
    """
    Endpoint para agregar una nueva sucursal al Excel
    """
    global branches_data
    
    try:
        branch_name = request.form.get('branch_name', '').strip()
        
        if not branch_name:
            return jsonify({
                'success': False,
                'message': 'Nombre de sucursal es requerido.'
            })
        
        if branches_data is None:
            branches_data = pd.DataFrame(columns=['Sucursal'])
        
        if branch_name in branches_data['Sucursal'].values:
            return jsonify({
                'success': False,
                'message': f'La sucursal "{branch_name}" ya existe.'
            })
        
        new_row = pd.DataFrame({'Sucursal': [branch_name]})
        branches_data = pd.concat([branches_data, new_row], ignore_index=True)
        
        # --- CORRECCIÓN ---
        # Se pasan el nombre de la hoja y el DataFrame a la función de guardado.
        if save_to_excel(BRANCHES_SHEET, branches_data):
            logger.info(f"Sucursal agregada exitosamente: {branch_name}")
            return jsonify({
                'success': True,
                'message': f'Sucursal "{branch_name}" agregada exitosamente.',
                'branch_name': branch_name
            })
        else:
            # Revertir el cambio en memoria si falla el guardado
            branches_data = branches_data[branches_data['Sucursal'] != branch_name]
            return jsonify({
                'success': False,
                'message': 'Error al guardar en el archivo Excel. Verifica que el archivo no esté abierto en otra aplicación.'
            })
        
    except Exception as e:
        logger.error(f"Error al agregar sucursal: {e}")
        return jsonify({
            'success': False,
            'message': f'Error interno del servidor: {str(e)}'
        })

        
@app.route('/get-excel-data', methods=['GET'])
def get_excel_data():
    """
    Lee todas las hojas de un archivo Excel y las devuelve como un solo JSON.
    Esta será la base de conocimiento para la IA.
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            logger.error(f"El archivo de base de conocimiento '{EXCEL_FILE}' no fue encontrado.")
            return jsonify({'success': False, 'message': 'Archivo de datos no encontrado.'}), 404

        xls = pd.ExcelFile(EXCEL_FILE)
        all_sheets_data = {}
        
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            
            # CORRECCIÓN JSON: Esta es una forma más robusta de manejar valores nulos (NaN).
            # Convierte el DataFrame a un string JSON válido, reemplazando NaN con null.
            # Luego, lo convierte de nuevo a un objeto de Python para que jsonify lo maneje correctamente.
            clean_json_string = df.to_json(orient='records', date_format='iso')
            all_sheets_data[sheet_name] = json.loads(clean_json_string)

        logger.info(f"Base de conocimiento cargada con {len(all_sheets_data)} hojas para la IA.")
        
        return jsonify({
            'success': True,
            'message': 'Datos cargados exitosamente.',
            'data': all_sheets_data
        })

    except Exception as e:
        logger.error(f"Error crítico al leer el archivo Excel para la IA: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor al leer los datos: {str(e)}'}), 500




@app.route('/delete-branch', methods=['POST'])
def delete_branch():
    """
    Endpoint para eliminar una sucursal del Excel
    """
    global branches_data
    
    try:
        branch_name = request.form.get('branch_name', '').strip()
        
        if not branch_name:
            return jsonify({
                'success': False,
                'message': 'Nombre de sucursal es requerido.'
            })
        
        # Verificar si los datos están disponibles
        if branches_data is None or branches_data.empty:
            return jsonify({
                'success': False,
                'message': 'Error del servidor: No se pudo cargar el archivo Excel.'
            })
        
        # Verificar si la sucursal existe
        if branch_name not in branches_data['Sucursal'].values:
            return jsonify({
                'success': False,
                'message': f'La sucursal "{branch_name}" no existe.'
            })
        
        # Eliminar la sucursal del DataFrame
        branches_data = branches_data[branches_data['Sucursal'] != branch_name]
        
        # --- LÍNEA CORREGIDA ---
        # Se pasan los parámetros correctos: el nombre de la hoja y el dataframe actualizado.
        if save_to_excel(BRANCHES_SHEET, branches_data):
            logger.info(f"Sucursal eliminada exitosamente: {branch_name}")
            return jsonify({
                'success': True,
                'message': f'Sucursal "{branch_name}" eliminada exitosamente.',
                'branch_name': branch_name
            })
        else:
            # Si falla, se recargan los datos para revertir el cambio en memoria
            load_all_data()
            return jsonify({
                'success': False,
                'message': 'Error al guardar en el archivo Excel. Los cambios han sido revertidos.'
            })
        
    except Exception as e:
        logger.error(f"Error al eliminar sucursal: {e}")
        return jsonify({
            'success': False,
            'message': f'Error interno del servidor: {str(e)}'
        })


@app.route('/egresos', methods=['POST'])
def handle_egresos():
    """Procesa el formulario de egresos y guarda los datos en la hoja 'Egresos'."""
    global egresos_data
    try:
        # Las claves del diccionario DEBEN estar en MAYÚSCULAS
        # para coincidir con la definición de EGRESOS_COLUMNS.
        form_data = {
            'FECHA': request.form.get('fecha', '').strip(),
            'COMPRA_SERVICIO': request.form.get('compra_servicio', '').strip(),
            'LUGAR_COMPRA': request.form.get('lugar_compra', '').strip(),
            'CONCEPTO': request.form.get('concepto', '').strip(),
            'SUCURSAL': request.form.get('sucursal', '').strip(),
            'OPERADOR': request.form.get('operador', '').strip(),
            'CANTIDAD': float(request.form.get('cantidad', '0')),
            'UDM': request.form.get('udm', '').strip(),
            'PRECIO': float(request.form.get('precio', '0')),
            'FECHA_REGISTRO': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        # Validación de campos requeridos
        required_fields = ['FECHA', 'COMPRA_SERVICIO', 'LUGAR_COMPRA', 'CONCEPTO', 'SUCURSAL', 'OPERADOR', 'UDM']
        if not all(form_data[field] for field in required_fields) or form_data['CANTIDAD'] <= 0 or form_data['PRECIO'] <= 0:
            return jsonify({'success': False, 'message': 'Todos los campos son requeridos y los valores numéricos deben ser mayores a cero.'})

        # Crear un nuevo registro y añadirlo al DataFrame
        nuevo_registro = pd.DataFrame([form_data])
        egresos_data = pd.concat([egresos_data, nuevo_registro], ignore_index=True)

        # --- CORRECCIÓN AQUÍ ---
        # Se especifica la hoja (EGRESOS_SHEET) y los datos (egresos_data) a guardar.
        if save_to_excel(EGRESOS_SHEET, egresos_data):
            logger.info(f"Egreso registrado exitosamente para el operador: {form_data['OPERADOR']}")
            return jsonify({'success': True, 'message': 'Egreso registrado exitosamente.'})
        else:
            egresos_data = egresos_data.iloc[:-1] # Revertir si falla el guardado
            return jsonify({'success': False, 'message': 'Error crítico al guardar en Excel.'})

    except ValueError as e:
        logger.error(f"Error de conversión de datos en el formulario de egresos: {e}")
        return jsonify({'success': False, 'message': 'Error en el formato de datos numéricos (Cantidad/Precio).'})
    except Exception as e:
        logger.error(f"Error al procesar el egreso: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {e}'})



        
@app.route('/egresos/list', methods=['GET'])
def list_egresos():
    """Lista todos los registros de egresos para el panel de gráficas."""
    if egresos_data is None or egresos_data.empty:
        return jsonify({'registros': [], 'total': 0})

    # Crear una copia para manipularla de forma segura
    data_copy = egresos_data.copy()

    # Pandas a menudo usa NaN para valores faltantes, lo que no es un JSON válido.
    # Lo reemplazamos por None, que se convierte en 'null' en JSON.
    data_copy = data_copy.where(pd.notnull(data_copy), None)

    # Convertir el DataFrame a una lista de diccionarios (JSON)
    registros = data_copy.to_dict('records')
    
    return jsonify({'registros': registros, 'total': len(registros)})

@app.route('/egresos/download', methods=['POST'])
def download_egresos():
    """
    Genera y envía un archivo Excel con los datos de egresos, con formato condicional.
    """
    try:
        # Recibe los datos de la tabla (incluyendo estado y observaciones) desde el frontend
        rows_data = request.get_json()
        if not rows_data:
            return jsonify({'success': False, 'message': 'No se recibieron datos para exportar.'}), 400

        # Crea una copia de los datos para no modificar el original mientras se renombra
        report_data = [row.copy() for row in rows_data]

        # Renombra las claves del JSON para que coincidan con las cabeceras del Excel
        for row in report_data:
            row['FECHA'] = row.pop('date', '')
            row['COMPRA'] = row.pop('compra', '')
            row['LUGAR DE COMPRA'] = row.pop('lugarCompra', '')
            row['CONCEPTO'] = row.pop('concepto', '')
            row['SUCURSAL'] = row.pop('sucursal', '')
            row['CANTIDAD'] = row.pop('cantidad', '')
            row['UDM'] = row.pop('udm', '')
            row['PRECIO'] = row.pop('precio', 0)
            row['COSTO X UDM'] = row.pop('costoUDM', '')
            row['OBSERVACIÓN'] = row.pop('observation', '')

        df = pd.DataFrame(report_data)

        # Define y ordena las columnas para el reporte final
        column_order = [
            'FECHA', 'COMPRA', 'LUGAR DE COMPRA', 'CONCEPTO', 'SUCURSAL',
            'CANTIDAD', 'UDM', 'PRECIO', 'COSTO X UDM', 'OBSERVACIÓN'
        ]
        # Asegura que todas las columnas existan
        for col in column_order:
            if col not in df.columns:
                df[col] = ''
        df = df[column_order]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Egresos')

            # Prepara los formatos de celda
            workbook = writer.book
            worksheet = writer.sheets['Egresos']
            red_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
            green_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
            currency_format = workbook.add_format({'num_format': '$#,##0'})
            
            # Formatos combinados para no perder el formato de moneda
            red_currency_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006', 'num_format': '$#,##0'})
            green_currency_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100', 'num_format': '$#,##0'})

            # Aplica el formato de moneda a la columna de Precio (H)
            worksheet.set_column('H:H', 12, currency_format)

            # Aplica el formato de color a las filas según el estado
            for idx, row_data in enumerate(rows_data):
                row_idx_excel = idx + 1  # Fila en Excel (1-indexado + cabecera)
                if row_data.get('status') == 'red':
                    worksheet.set_row(row_idx_excel, None, red_format)
                    worksheet.write_number(row_idx_excel, 7, row_data.get('precio', 0), red_currency_format)
                elif row_data.get('status') == 'green':
                    worksheet.set_row(row_idx_excel, None, green_format)
                    worksheet.write_number(row_idx_excel, 7, row_data.get('precio', 0), green_currency_format)

            # Ajusta el ancho de las columnas
            for i, col in enumerate(df.columns):
                width = max(df[col].astype(str).map(len).max(), len(col))
                worksheet.set_column(i, i, width + 2)

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='reporte_egresos_formateado.xlsx'
        )
    except Exception as e:
        logger.error(f"Error al generar el archivo Excel de egresos: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {e}'}), 500


@app.route('/nomina/download', methods=['POST'])
def download_nomina():
    """
    Genera y envía un archivo Excel con los datos de nómina, con formato condicional.
    """
    try:
        # Recibe los datos de la tabla (incluyendo estado y observaciones) desde el frontend
        rows_data = request.get_json()
        if not rows_data:
            return jsonify({'success': False, 'message': 'No se recibieron datos para exportar.'}), 400

        # Crea una copia de los datos para manipularlos
        report_data = [row.copy() for row in rows_data]

        # Renombra las claves del JSON para que coincidan con las cabeceras del Excel
        for row in report_data:
            row['FECHA'] = row.pop('date', '')
            row['COLABORADOR'] = row.pop('colaborador', '')
            row['DIAS LABORADOS'] = row.pop('diasLaborados', '')
            row['SUCURSAL'] = row.pop('sucursal', '')
            row['CANTIDAD'] = row.pop('cantidad', 0)
            row['TIPO NOMINA'] = row.pop('nomina', '')
            row['NOTAS'] = row.pop('notas', '')
            row['OBSERVACIÓN'] = row.pop('observation', '')
        
        df = pd.DataFrame(report_data)

        # Define y ordena las columnas para el reporte final
        column_order = [
            'FECHA', 'COLABORADOR', 'DIAS LABORADOS', 'SUCURSAL', 
            'CANTIDAD', 'TIPO NOMINA', 'NOTAS', 'OBSERVACIÓN'
        ]
        # Asegura que todas las columnas existan
        for col in column_order:
            if col not in df.columns:
                df[col] = ''
        df = df[column_order]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Nomina')

            # Prepara los formatos de celda
            workbook = writer.book
            worksheet = writer.sheets['Nomina']
            red_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
            green_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
            currency_format = workbook.add_format({'num_format': '$#,##0.00'})
            
            # Formatos combinados para no perder el formato de moneda
            red_currency_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006', 'num_format': '$#,##0.00'})
            green_currency_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100', 'num_format': '$#,##0.00'})

            # Aplica el formato de moneda a la columna de Cantidad (E)
            worksheet.set_column('E:E', 15, currency_format)

            # Aplica el formato de color a las filas según el estado
            for idx, row_data in enumerate(rows_data):
                row_idx_excel = idx + 1  # Fila en Excel (1-indexado + cabecera)
                if row_data.get('status') == 'red':
                    worksheet.set_row(row_idx_excel, None, red_format)
                    # Reescribe la celda de cantidad con el formato combinado
                    worksheet.write_number(row_idx_excel, 4, row_data.get('cantidad', 0), red_currency_format)
                elif row_data.get('status') == 'green':
                    worksheet.set_row(row_idx_excel, None, green_format)
                    # Reescribe la celda de cantidad con el formato combinado
                    worksheet.write_number(row_idx_excel, 4, row_data.get('cantidad', 0), green_currency_format)

            # Ajusta el ancho de las columnas
            for i, col in enumerate(df.columns):
                width = max(df[col].astype(str).map(len).max(), len(col))
                worksheet.set_column(i, i, width + 2)

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='reporte_nomina_formateado.xlsx'
        )
    except Exception as e:
        logger.error(f"Error al generar el archivo Excel de nómina: {e}")
        return jsonify({'success': False, 'message': f'Error interno del servidor: {e}'}), 500


@app.route('/reload-data', methods=['POST'])
def reload_data():
    """
    Endpoint para recargar todas las hojas desde el archivo Excel
    """
    # --- CORRECCIÓN CLAVE: Declarar todas las variables globales que se van a modificar ---
    global user_data, admin_data, branches_data, cierre_caja_data, egresos_data, nomina_data, registro_data, permisos_data, config_data
    
    try:
        # --- CORRECCIÓN CLAVE: Desempaquetar correctamente TODAS las tablas de datos ---
        user_data, admin_data, branches_data, cierre_caja_data, egresos_data, nomina_data, registro_data, permisos_data, config_data = load_all_data()
        
        users_count = len(user_data) if user_data is not None else 0
        admins_count = len(admin_data) if admin_data is not None else 0
        branches_count = len(branches_data) if branches_data is not None else 0
        cierre_count = len(cierre_caja_data) if cierre_caja_data is not None else 0
        # Puedes añadir más contadores si lo deseas para la respuesta
        
        return jsonify({
            'success': True,
            'message': f'Datos recargados exitosamente.',
            'details': {
                'usuarios': f'{users_count} registros cargados',
                'administradores': f'{admins_count} registros cargados',
                'sucursales': f'{branches_count} registros cargados',
                'cierre_caja': f'{cierre_count} registros cargados',
                'registros_asistencia': f'{len(registro_data) if registro_data is not None else 0} registros cargados'
            }
        })
        
    except Exception as e:
        logger.error(f"Error al recargar datos: {e}")
        return jsonify({
            'success': False,
            'message': f'Error al recargar datos: {str(e)}'
        })


@app.route('/employees', methods=['GET'])
def get_employees():
    """Devuelve la lista de empleados con los campos que el Checkador.html espera."""
    if user_data.empty:
        return jsonify([])
    
    # Renombrar columnas para que coincidan con el frontend
    employees_df = user_data.rename(columns={
        'IDHuella': 'EmployeeID',
        'Usuario': 'Name',
        'Phone': 'Phone',
        'InternalID': 'InternalID',
        'WorkArea': 'WorkArea',
        'CustomEntryTime': 'CustomEntryTime',
        'CustomExitTime': 'CustomExitTime'
    })
    
    # Seleccionar solo las columnas necesarias para el frontend
    display_columns = ['EmployeeID', 'Name', 'Phone', 'InternalID', 'WorkArea', 'CustomEntryTime', 'CustomExitTime']
    employees_df = employees_df[display_columns]

    # Reemplazar 'nan' o valores vacíos por strings vacíos
    employees_list = employees_df.fillna('').to_dict('records')
    return jsonify(employees_list)

# app.py

@app.route('/employee/save', methods=['POST'])
def save_employee():
    """Crea o actualiza un empleado en la hoja 'Usuarios'."""
    global user_data
    try:
        form_data = request.form
        # ID de Huella nuevo o actual
        employee_id = form_data.get('employeeID')
        # ID de Huella original, para buscar al usuario si este se cambia
        original_id = form_data.get('originalEmployeeID', employee_id)

        if not employee_id:
            return jsonify({'success': False, 'message': 'El ID de Huella es obligatorio.'}), 400

        # Busca al usuario por su ID original.
        existing_user_idx = user_data[user_data['IDHuella'] == original_id].index

        # Datos a guardar del formulario
        update_data = {
            'IDHuella': employee_id,
            'Usuario': form_data.get('name'),
            'CustomEntryTime': form_data.get('customEntryTime'),
            'CustomExitTime': form_data.get('customExitTime')
        }

        if not existing_user_idx.empty:  # ----- ACTUALIZAR EMPLEADO EXISTENTE -----
            idx = existing_user_idx[0]
            # Actualiza cada campo en el DataFrame de pandas
            for key, value in update_data.items():
                if key in user_data.columns:
                    user_data.loc[idx, key] = value
            message = 'Empleado actualizado correctamente.'

        else:  # ----- CREAR NUEVO EMPLEADO -----
            # Para un nuevo empleado, llenamos los campos esenciales
            new_employee_row = {
                'Usuario': form_data.get('name'),
                'Contraseña': 'SinAsignar',  # Se puede poner una contraseña por defecto
                'Sucursal': 'SinAsignar',    # O una sucursal por defecto
                'IDHuella': employee_id,
                'Phone': '',
                'InternalID': '',
                'WorkArea': '',
                'CustomEntryTime': form_data.get('customEntryTime'),
                'CustomExitTime': form_data.get('customExitTime')
            }
            new_df = pd.DataFrame([new_employee_row])
            user_data = pd.concat([user_data, new_df], ignore_index=True)
            message = 'Empleado creado correctamente.'

        # Guardar los cambios en el archivo Excel
        if save_to_excel(USERS_SHEET, user_data):
            return jsonify({'success': True, 'message': message})
        else:
            # Si falla, recargar desde el archivo para descartar cambios en memoria
            load_all_data()
            return jsonify({'success': False, 'message': 'Error al guardar en Excel. Cambios revertidos.'}), 500

    except Exception as e:
        logger.error(f"Error en /employee/save: {e}")
        return jsonify({'success': False, 'message': f'Error interno: {e}'}), 500

@app.route('/config/save', methods=['POST'])
def save_config():
    """Guarda la configuración en la hoja 'Config'."""
    global config_data
    try:
        new_config = request.form.to_dict()
        config_data = pd.DataFrame(list(new_config.items()), columns=['key', 'value'])
        
        # --- LÍNEA CORREGIDA ---
        if save_to_excel(CONFIG_SHEET, config_data):
            return jsonify({'success': True, 'message': 'Configuración guardada.'})
        else:
            load_all_data()
            return jsonify({'success': False, 'message': 'Error al guardar la configuración. Cambios revertidos.'}), 500
    except Exception as e:
        logger.error(f"Error en /config/save: {e}")
        return jsonify({'success': False, 'message': f'Error interno: {e}'}), 500

@app.route('/config', methods=['GET'])
def get_config():
    """Obtiene la configuración desde la hoja 'Config'."""
    if config_data.empty:
        # Valores por defecto si no hay configuración
        return jsonify({
            'standard_entry_time': '09:00',
            'standard_exit_time': '18:00',
            'late_threshold_minutes': 10,
            'early_leave_threshold_minutes': 10
        })
    
    # Convierte el DataFrame de config a un diccionario
    config_dict = pd.Series(config_data.value.values, index=config_data.key).to_dict()
    return jsonify(config_dict)

@app.route('/attendance-log', methods=['GET'])
def get_attendance_log():
    """Devuelve los registros de entrada/salida para una fecha específica."""
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'error': 'Falta el parámetro de fecha'}), 400

    # Filtra los registros por el Dia solicitado
    daily_logs = registro_data[registro_data['Dia'] == date_str]
    
    if daily_logs.empty:
        return jsonify([])

    # Simular la estructura que espera el frontend (con múltiples entradas/salidas)
    # y añadir el nombre del empleado
    result_logs = []
    for _, row in daily_logs.iterrows():
        user_info = user_data[user_data['Usuario'] == row['Usuario']]
        name = row['Usuario']
        employee_id = user_info['IDHuella'].iloc[0] if not user_info.empty else 'N/A'

        if pd.notna(row['HoraEntrada']) and row['HoraEntrada']:
            result_logs.append({
                'EmployeeID': employee_id,
                'Name': name,
                'Timestamp': f"{row['Dia']}T{row['HoraEntrada']}",
                'Action': 'Entrada'
            })
        if pd.notna(row['HoraSalida']) and row['HoraSalida']:
            result_logs.append({
                'EmployeeID': employee_id,
                'Name': name,
                'Timestamp': f"{row['Dia']}T{row['HoraSalida']}",
                'Action': 'Salida'
            })
            
    return jsonify(result_logs)

@app.route('/attendance-log-range', methods=['GET'])
def get_attendance_log_range():
    """Devuelve registros de asistencia en un rango de fechas."""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        return jsonify({'error': 'Faltan parámetros de fecha de inicio o fin'}), 400

    # Filtra los registros que están dentro del rango de fechas
    range_logs = registro_data[
        (registro_data['Dia'] >= start_date) & 
        (registro_data['Dia'] <= end_date)
    ]

    # Añadir el nombre del empleado a cada registro
    merged_df = pd.merge(range_logs, user_data[['Usuario', 'IDHuella']], on='Usuario', how='left')
    merged_df.rename(columns={'IDHuella': 'EmployeeID', 'Usuario': 'Name', 'Dia': 'Date'}, inplace=True)
    
    # Simular la estructura que espera el frontend
    result_logs = []
    for _, row in merged_df.iterrows():
        if pd.notna(row['HoraEntrada']) and row['HoraEntrada']:
            result_logs.append({
                'EmployeeID': row['EmployeeID'], 'Name': row['Name'], 'Date': row['Date'],
                'Timestamp': f"{row['Date']}T{row['HoraEntrada']}", 'Action': 'Entrada'
            })
        if pd.notna(row['HoraSalida']) and row['HoraSalida']:
             result_logs.append({
                'EmployeeID': row['EmployeeID'], 'Name': row['Name'], 'Date': row['Date'],
                'Timestamp': f"{row['Date']}T{row['HoraSalida']}", 'Action': 'Salida'
            })

    return jsonify(result_logs)

@app.route('/permissions', methods=['GET'])
def get_permissions():
    """Devuelve la lista de todos los permisos y vacaciones."""
    if permisos_data.empty:
        return jsonify([])
    # Renombrar columnas para el frontend
    perms_df = permisos_data.rename(columns={
        'Empleado': 'EmployeeID', 'Tipo': 'Type',
        'FechaInicio': 'StartDate', 'FechaFinal': 'EndDate',
        'Descripcion': 'Description'
    })
    return jsonify(perms_df.fillna('').to_dict('records'))

@app.route('/status')
def status():
    """
    Endpoint para verificar el estado completo del servidor
    """
    users_count = len(user_data) if user_data is not None else 0
    admins_count = len(admin_data) if admin_data is not None else 0
    branches_count = len(branches_data) if branches_data is not None else 0
    cierre_count = len(cierre_caja_data) if cierre_caja_data is not None else 0
    
    return jsonify({
        'server': 'Sushi Tlalpan Login Server',
        'status': 'running',
        'excel_file': EXCEL_FILE,
        'file_exists': os.path.exists(EXCEL_FILE),
        'sheets': {
            'usuarios': {
                'sheet_name': USERS_SHEET,
                'records_loaded': users_count,
                'status': 'ok' if user_data is not None else 'error'
            },
            'administradores': {
                'sheet_name': ADMINS_SHEET,
                'records_loaded': admins_count,
                'status': 'ok' if admin_data is not None else 'error'
            },
            'sucursales': {
                'sheet_name': BRANCHES_SHEET,
                'records_loaded': branches_count,
                'status': 'ok' if branches_data is not None else 'error'
            },
            'cierre_caja': {
                'sheet_name': CIERRE_CAJA_SHEET,
                'records_loaded': cierre_count,
                'status': 'ok' if cierre_caja_data is not None else 'error',
                'columns': CIERRE_CAJA_COLUMNS
            }
        },
        'total_records': users_count + admins_count + branches_count + cierre_count
    })

@app.route('/permission/save', methods=['POST'])
def save_permission():
    """Guarda un nuevo permiso o vacación."""
    global permisos_data
    try:
        form = request.form
        new_permission = pd.DataFrame([{
            'IDPermiso': str(int(datetime.now().timestamp())), # ID único
            'Empleado': form.get('employeeID'),
            'Tipo': form.get('type'),
            'FechaInicio': form.get('startDate'),
            'FechaFinal': form.get('endDate'),
            'Descripcion': form.get('description')
        }])
        
        permisos_data = pd.concat([permisos_data, new_permission], ignore_index=True)
        
        # --- CORRECCIÓN AQUÍ ---
        # Se especifica la hoja (PERMISOS_SHEET) y los datos (permisos_data) a guardar.
        if save_to_excel(PERMISOS_SHEET, permisos_data):
            return jsonify({'success': True, 'message': 'Permiso guardado correctamente.'})
        else:
            # Si falla, revertimos el cambio en memoria
            permisos_data = permisos_data.iloc[:-1]
            return jsonify({'success': False, 'message': 'Error al guardar el permiso en Excel.'})
            
    except Exception as e:
        logger.error(f"Error en /permission/save: {e}")
        return jsonify({'success': False, 'message': f'Error interno: {e}'})

# --- Manejo de Errores ---
@app.errorhandler(404)
def not_found(error):
    return jsonify({
        'error': 'Endpoint no encontrado',
        'message': 'La ruta solicitada no existe',
        'available_endpoints': [
            '/login', 
            '/admin-login', 
            '/status', 
            '/users', 
            '/admins', 
            '/branches',
            '/cierre-caja',
            '/cierre-caja/list',
            '/cierre-caja/search'
        ]
    }), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({
        'error': 'Error interno del servidor',
        'message': 'Ocurrió un error inesperado'
    }), 500

if __name__ == '__main__':
    print("="*60)
    print("SERVIDOR SUSHI TLALPAN - VERSIÓN DEFINITIVA")
    print("="*60)
    print(f"Archivo Excel: {EXCEL_FILE}")
    print("Servidor corriendo en: http://localhost:8080")
    print("="*60)
    app.run(host='0.0.0.0', port=8080, debug=False, use_reloader=False)
