"""
Excel Import Endpoints

This module provides endpoints for importing data from Excel files into the database.
Supports importing Sales (Ventas), Expenses (Egresos), and Payroll (Nomina) records.
"""

from typing import List, Dict, Any, Optional
from uuid import UUID
from datetime import datetime, date, timezone
import io

from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, status
from sqlalchemy.orm import Session
import pandas as pd
from openpyxl import load_workbook

from backend.fastapi.dependencies.database import get_sync_db
from backend.fastapi.models.admin import Admin
from backend.fastapi.models.user import User
from backend.fastapi.models.branch import Branch
from backend.fastapi.models.sales import Sales
from backend.fastapi.models.expense import Expense
from backend.fastapi.models.payroll import Payroll
from backend.fastapi.core.utils import normalize_username, normalize_branch_name
from backend.security.dependencies import RequireActiveAdmin


router = APIRouter(prefix="/import", tags=["Excel Import"])


def build_lookup_maps(db: Session) -> tuple[Dict[str, UUID], Dict[str, UUID]]:
    """
    Build lookup dictionaries for users and branches with normalized keys.
    
    Returns:
        Tuple of (users_map, branches_map) where keys are normalized names
    """
    # Load all non-deleted users
    users = db.query(User).filter(User.deleted_at.is_(None)).all()
    users_map = {normalize_username(user.username): user.id for user in users}
    
    # Load all non-deleted branches
    branches = db.query(Branch).filter(Branch.deleted_at.is_(None)).all()
    branches_map = {normalize_branch_name(branch.name): branch.id for branch in branches}
    
    return users_map, branches_map


def parse_date(date_value: Any) -> Optional[date]:
    """
    Parse various date formats from Excel into a date object.
    Handles formats like: 2025-07-05, datetime objects, Excel serial dates
    """
    if pd.isna(date_value) or date_value is None or date_value == '':
        return None
    
    try:
        # If it's already a datetime/date object
        if isinstance(date_value, (datetime, date)):
            return date_value if isinstance(date_value, date) else date_value.date()
        
        # Try to parse as string (handles ISO format like "2025-07-05")
        parsed = pd.to_datetime(date_value, errors='coerce')
        if pd.isna(parsed):
            return None
        return parsed.date()
    except Exception:
        return None


def parse_datetime(datetime_value: Any) -> Optional[datetime]:
    """
    Parse various datetime formats from Excel into a datetime object.
    Handles formats like: 2025-07-10 21:39:06, datetime objects, Excel serial dates
    """
    if pd.isna(datetime_value) or datetime_value is None or datetime_value == '':
        return None
    
    try:
        # If it's already a datetime object
        if isinstance(datetime_value, datetime):
            # Ensure it has timezone info (UTC)
            if datetime_value.tzinfo is None:
                return datetime_value.replace(tzinfo=timezone.utc)
            return datetime_value
        
        # If it's a date object, convert to datetime
        if isinstance(datetime_value, date):
            return datetime.combine(datetime_value, datetime.min.time()).replace(tzinfo=timezone.utc)
        
        # Try to parse as string (handles formats like "2025-07-10 21:39:06")
        parsed = pd.to_datetime(datetime_value, errors='coerce')
        if pd.isna(parsed):
            return None
        
        # Convert to timezone-aware datetime (UTC)
        if parsed.tzinfo is None:
            return parsed.replace(tzinfo=timezone.utc)
        return parsed
    except Exception:
        return None


def safe_float(value: Any, default: float = 0.0, allow_none: bool = True) -> Optional[float]:
    """
    Safely convert value to float.
    
    Args:
        value: Value to convert
        default: Default value if conversion fails
        allow_none: If False, will return None for invalid values instead of default
        
    Returns:
        Float value, default, or None
        
    Rejects:
        - Excel errors (#VALUE!, #REF!, #DIV/0!, etc.)
        - String literals like 'None', 'N/A', 'null'
    """
    if pd.isna(value) or value is None or value == '':
        return default
    
    # Check for Excel error values
    if isinstance(value, str):
        value_upper = value.strip().upper()
        # Excel error values
        if value_upper.startswith('#') or value_upper in ['NONE', 'N/A', 'NULL', 'NA']:
            return None if not allow_none else default
    
    try:
        return float(value)
    except (ValueError, TypeError):
        return None if not allow_none else default


def safe_int(value: Any, default: int = 0) -> int:
    """Safely convert value to int."""
    if pd.isna(value) or value is None or value == '':
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def safe_str(value: Any) -> str:
    """Safely convert value to string."""
    if pd.isna(value) or value is None:
        return ""
    return str(value).strip()


@router.post("/sales", summary="Import Sales from Excel")
async def import_sales_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_sync_db),
    current_admin: Admin = RequireActiveAdmin
):
    """
    Import sales records from an Excel file (Ventas sheet).
    
    **Required Excel Columns:**
    - OPERADOR: Username of the worker
    - SUCURSAL: Branch name
    - FECHA DE CORTE: Closure date
    - NO. CIERRE: Closure number
    - VENTAS TOTALES: Total sales amount
    - TARJETA SEGÚN ITPV: Card amount according to ITPV
    - DEVOLUCIÓN EN TARJETA: Card refunds
    - TARJETA/ KIWI: Kiwi card payments
    - TRANSFER: Bank transfer amounts
    - EFECTIVO: Cash amount
    - DEVOLUCIÓN EN EFECTIVO: Cash refunds
    - NO. PAGOS: Number of payments
    - TOTAL COMISIÓN KIWI: Total Kiwi fees
    - NOTAS: Notes (optional)
    - FECHA_REGISTRO: Registration date (optional, defaults to current time)
    
    **Auto-calculated fields:**
    - TOTAL TARJETA REAL: Calculated from card amounts
    - TOTAL EN EFECTIVO REAL: Calculated from cash amounts
    - DIFERENCIA REVISAR: Calculated discrepancy
    - TICKET PROMEDIO: Average sale amount
    - TOTAL SIN COMISIÓN: Kiwi amount minus fees
    - TOTAL DE INGRESOS: Final revenue total
    
    **Process:**
    1. Validates all rows before importing
    2. Skips rows with invalid users or branches
    3. Calculates all derived fields automatically
    4. Uses database transaction (all or nothing per file)
    5. Returns detailed report of imported vs skipped rows
    
    **Permissions:** Requires active admin authentication
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be an Excel file (.xlsx or .xls)"
        )
    
    try:
        # Read Excel file
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        
        # Check if Ventas sheet exists
        if 'Ventas' not in wb.sheetnames:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must contain a 'Ventas' sheet"
            )
        
        # Read Ventas sheet
        df = pd.read_excel(io.BytesIO(contents), sheet_name='Ventas')
        
        if df.empty:
            return {
                "success": True,
                "message": "No records found in Ventas sheet",
                "total_rows": 0,
                "imported": 0,
                "skipped": 0,
                "skipped_rows": []
            }
        
        # Build lookup maps
        users_map, branches_map = build_lookup_maps(db)
        
        # Validate and prepare records
        valid_records = []
        skipped_rows = []
        
        for idx, row in df.iterrows():
            row_num = idx + 2  # Excel row number (1-indexed + header)
            errors = []
            
            # Extract and normalize identifiers
            operador = safe_str(row.get('OPERADOR', ''))
            sucursal = safe_str(row.get('SUCURSAL', ''))
            
            if not operador:
                errors.append("Missing OPERADOR")
            if not sucursal:
                errors.append("Missing SUCURSAL")
            
            # Lookup user and branch
            worker_id = None
            branch_id = None
            
            if operador:
                normalized_user = normalize_username(operador)
                worker_id = users_map.get(normalized_user)
                if not worker_id:
                    errors.append(f"User '{operador}' not found in database")
            
            if sucursal:
                normalized_branch = normalize_branch_name(sucursal)
                branch_id = branches_map.get(normalized_branch)
                if not branch_id:
                    errors.append(f"Branch '{sucursal}' not found in database")
            
            # Parse other fields
            closure_date = parse_date(row.get('FECHA DE CORTE'))
            if not closure_date:
                errors.append("Missing or invalid FECHA DE CORTE")
            
            closure_number = safe_int(row.get('NO. CIERRE'))
            if not closure_number:
                errors.append("Missing NO. CIERRE")
            
            # Parse all sales columns with strict validation
            sales_total = safe_float(row.get('VENTAS TOTALES', 0))
            card_itpv = safe_float(row.get('TARJETA SEGÚN ITPV', 0))
            card_refund = safe_float(row.get('DEVOLUCIÓN EN TARJETA', 0))
            
            # Validate TARJETA/ KIWI (reject None, #VALUE!, etc.)
            card_kiwi = safe_float(row.get('TARJETA/ KIWI', 0), allow_none=False)
            if card_kiwi is None:
                errors.append("Invalid value in TARJETA/ KIWI (found 'None' or Excel error)")
                card_kiwi = 0  # Set to 0 for display purposes
            
            # Validate TRANSFER (reject None, #VALUE!, etc.)
            transfer_amt = safe_float(row.get('TRANSFER', 0), allow_none=False)
            if transfer_amt is None:
                errors.append("Invalid value in TRANSFER (found 'None' or Excel error)")
                transfer_amt = 0  # Set to 0 for display purposes
            
            cash_amt = safe_float(row.get('EFECTIVO', 0))
            cash_refund = safe_float(row.get('DEVOLUCIÓN EN EFECTIVO', 0))
            payments_nbr = safe_int(row.get('NO. PAGOS', 0))
            kiwi_fee_total = safe_float(row.get('TOTAL COMISIÓN KIWI', 0))
            notes = safe_str(row.get('NOTAS', ''))
            
            # Validate numeric fields for negative values
            if card_itpv < 0:
                errors.append("TARJETA SEGÚN ITPV cannot be negative")
            if card_refund < 0:
                errors.append("DEVOLUCIÓN EN TARJETA cannot be negative")
            if cash_amt < 0:
                errors.append("EFECTIVO cannot be negative")
            if cash_refund < 0:
                errors.append("DEVOLUCIÓN EN EFECTIVO cannot be negative")
            if kiwi_fee_total < 0:
                errors.append("TOTAL COMISIÓN KIWI cannot be negative")
            
            # Parse FECHA_REGISTRO if available (otherwise use current time)
            fecha_registro = parse_datetime(row.get('FECHA_REGISTRO'))
            
            # If there are errors, skip this row
            if errors:
                skipped_rows.append({
                    "row": row_num,
                    "operador": operador,
                    "sucursal": sucursal,
                    "errors": errors
                })
                continue
            
            # Add valid record
            valid_records.append({
                "worker_id": worker_id,
                "branch_id": branch_id,
                "closure_date": closure_date,
                "closure_number": closure_number,
                "payments_nbr": payments_nbr,
                "sales_total": sales_total,
                "card_itpv": card_itpv,
                "card_refund": card_refund,
                "card_kiwi": card_kiwi,
                "transfer_amt": transfer_amt,
                "cash_amt": cash_amt,
                "cash_refund": cash_refund,
                "kiwi_fee_total": kiwi_fee_total,
                "notes": notes if notes else None,
                "created_at": fecha_registro if fecha_registro else datetime.now(timezone.utc),
                "review_state": "pending"
            })
        
        # If no valid records, return early
        if not valid_records:
            return {
                "success": False,
                "message": "No valid records to import",
                "total_rows": len(df),
                "imported": 0,
                "skipped": len(skipped_rows),
                "skipped_rows": skipped_rows
            }
        
        # Import valid records in a transaction
        try:
            imported_count = 0
            for record_data in valid_records:
                sales_record = Sales(**record_data)
                # Calculate derived fields (card_total, cash_total, discrepancy, avg_sale, etc.)
                sales_record.calculate_totals()
                db.add(sales_record)
                imported_count += 1
            
            db.commit()
            
            return {
                "success": True,
                "message": f"Successfully imported {imported_count} sales records",
                "total_rows": len(df),
                "imported": imported_count,
                "skipped": len(skipped_rows),
                "skipped_rows": skipped_rows
            }
            
        except Exception as e:
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Database error during import: {str(e)}"
            )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing Excel file: {str(e)}"
        )


@router.post("/expenses", summary="Import Expenses from Excel")
async def import_expenses_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_sync_db),
    current_admin: Admin = RequireActiveAdmin
):
    """
    Import expense records from an Excel file (Egresos sheet).
    
    **Required Excel Columns:**
    - OPERADOR: Username of the worker
    - SUCURSAL: Branch name
    - FECHA: Expense date
    - COMPRA_SERVICIO: Expense description (maps to expense_description)
    - LUGAR_COMPRA: Vendor/payee name (maps to vendor_payee)
    - CONCEPTO: Expense category (maps to expense_category)
    - CANTIDAD: Quantity purchased (maps to quantity)
    - UDM: Unit of measure (maps to unit_of_measure, defaults to 'each')
    - PRECIO: Unit price (maps to unit_cost)
    - FECHA_REGISTRO: Registration date (optional, defaults to current time)
    
    **Auto-calculated fields:**
    - total_amount: CANTIDAD * PRECIO
    
    **Process:**
    1. Validates all rows before importing
    2. Skips rows with invalid users or branches
    3. Calculates total_amount and validates unit_cost
    4. Uses database transaction (all or nothing per file)
    5. Returns detailed report of imported vs skipped rows
    
    **Permissions:** Requires active admin authentication
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be an Excel file (.xlsx or .xls)"
        )
    
    try:
        # Read Excel file
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        
        # Check if Egresos sheet exists
        if 'Egresos' not in wb.sheetnames:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must contain an 'Egresos' sheet"
            )
        
        # Read Egresos sheet
        df = pd.read_excel(io.BytesIO(contents), sheet_name='Egresos')
        
        if df.empty:
            return {
                "success": True,
                "message": "No records found in Egresos sheet",
                "total_rows": 0,
                "imported": 0,
                "skipped": 0,
                "skipped_rows": []
            }
        
        # Build lookup maps
        users_map, branches_map = build_lookup_maps(db)
        
        # Validate and prepare records
        valid_records = []
        skipped_rows = []
        
        for idx, row in df.iterrows():
            row_num = idx + 2  # Excel row number (1-indexed + header)
            errors = []
            
            # Extract and normalize identifiers
            operador = safe_str(row.get('OPERADOR', ''))
            sucursal = safe_str(row.get('SUCURSAL', ''))
            
            if not operador:
                errors.append("Missing OPERADOR")
            if not sucursal:
                errors.append("Missing SUCURSAL")
            
            # Lookup user and branch
            worker_id = None
            branch_id = None
            
            if operador:
                normalized_user = normalize_username(operador)
                worker_id = users_map.get(normalized_user)
                if not worker_id:
                    errors.append(f"User '{operador}' not found in database")
            
            if sucursal:
                normalized_branch = normalize_branch_name(sucursal)
                branch_id = branches_map.get(normalized_branch)
                if not branch_id:
                    errors.append(f"Branch '{sucursal}' not found in database")
            
            # Parse other fields
            expense_date = parse_date(row.get('FECHA'))
            if not expense_date:
                errors.append("Missing or invalid FECHA")
            
            expense_description = safe_str(row.get('COMPRA_SERVICIO'))
            if not expense_description:
                errors.append("Missing COMPRA_SERVICIO")
            
            quantity = safe_float(row.get('CANTIDAD', 1))
            unit_cost = safe_float(row.get('PRECIO', 0))
            total_amount = quantity * unit_cost
            
            if total_amount <= 0:
                errors.append("CANTIDAD * PRECIO must be greater than 0")
            
            vendor_payee = safe_str(row.get('LUGAR_COMPRA', ''))
            expense_category = safe_str(row.get('CONCEPTO', ''))
            unit_of_measure = safe_str(row.get('UDM', 'each'))
            
            # Parse FECHA_REGISTRO if available (otherwise use current time)
            fecha_registro = parse_datetime(row.get('FECHA_REGISTRO'))
            
            # If there are errors, skip this row
            if errors:
                skipped_rows.append({
                    "row": row_num,
                    "operador": operador,
                    "sucursal": sucursal,
                    "errors": errors
                })
                continue
            
            # Add valid record
            valid_records.append({
                "worker_id": worker_id,
                "branch_id": branch_id,
                "expense_date": expense_date,
                "expense_description": expense_description,
                "quantity": quantity,
                "unit_cost": unit_cost,
                "total_amount": total_amount,
                "unit_of_measure": unit_of_measure if unit_of_measure else "each",
                "vendor_payee": vendor_payee if vendor_payee else None,
                "expense_category": expense_category if expense_category else None,
                "created_at": fecha_registro if fecha_registro else datetime.now(timezone.utc),
                "review_state": "pending"
            })
        
        # If no valid records, return early
        if not valid_records:
            return {
                "success": False,
                "message": "No valid records to import",
                "total_rows": len(df),
                "imported": 0,
                "skipped": len(skipped_rows),
                "skipped_rows": skipped_rows
            }
        
        # Import valid records in a transaction
        try:
            imported_count = 0
            for record_data in valid_records:
                expense_record = Expense(**record_data)
                # Calculate unit_cost if needed (validates consistency)
                expense_record.calculate_unit_cost()
                db.add(expense_record)
                imported_count += 1
            
            db.commit()
            
            return {
                "success": True,
                "message": f"Successfully imported {imported_count} expense records",
                "total_rows": len(df),
                "imported": imported_count,
                "skipped": len(skipped_rows),
                "skipped_rows": skipped_rows
            }
            
        except Exception as e:
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Database error during import: {str(e)}"
            )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing Excel file: {str(e)}"
        )


@router.post("/payroll", summary="Import Payroll from Excel")
async def import_payroll_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_sync_db),
    current_admin: Admin = RequireActiveAdmin
):
    """
    Import payroll records from an Excel file (Nomina sheet).
    
    **Required Excel Columns:**
    - OPERADOR: Username of the worker
    - SUCURSAL: Branch name
    - FECHA: Payroll date
    - DIAS_LABORADOS: Days worked
    - CANTIDAD: Payment amount
    - TIPO_NOMINA: Type of payroll (regular, overtime, bonus, etc.) - defaults to 'regular'
    - NOTAS: Notes (optional)
    - FECHA_REGISTRO: Registration date (optional, defaults to current time)
    
    **Process:**
    1. Validates all rows before importing
    2. Skips rows with invalid users or branches
    3. Uses database transaction (all or nothing per file)
    4. Returns detailed report of imported vs skipped rows
    
    **Permissions:** Requires active admin authentication
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be an Excel file (.xlsx or .xls)"
        )
    
    try:
        # Read Excel file
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        
        # Check if Nomina sheet exists
        if 'Nomina' not in wb.sheetnames:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must contain a 'Nomina' sheet"
            )
        
        # Read Nomina sheet
        df = pd.read_excel(io.BytesIO(contents), sheet_name='Nomina')
        
        if df.empty:
            return {
                "success": True,
                "message": "No records found in Nomina sheet",
                "total_rows": 0,
                "imported": 0,
                "skipped": 0,
                "skipped_rows": []
            }
        
        # Build lookup maps
        users_map, branches_map = build_lookup_maps(db)
        
        # Validate and prepare records
        valid_records = []
        skipped_rows = []
        
        for idx, row in df.iterrows():
            row_num = idx + 2  # Excel row number (1-indexed + header)
            errors = []
            
            # Extract and normalize identifiers
            operador = safe_str(row.get('OPERADOR', ''))
            sucursal = safe_str(row.get('SUCURSAL', ''))
            
            if not operador:
                errors.append("Missing OPERADOR")
            if not sucursal:
                errors.append("Missing SUCURSAL")
            
            # Lookup user and branch
            worker_id = None
            branch_id = None
            
            if operador:
                normalized_user = normalize_username(operador)
                worker_id = users_map.get(normalized_user)
                if not worker_id:
                    errors.append(f"User '{operador}' not found in database")
            
            if sucursal:
                normalized_branch = normalize_branch_name(sucursal)
                branch_id = branches_map.get(normalized_branch)
                if not branch_id:
                    errors.append(f"Branch '{sucursal}' not found in database")
            
            # Parse other fields
            payroll_date = parse_date(row.get('FECHA'))
            if not payroll_date:
                errors.append("Missing or invalid FECHA")
            
            days_worked = safe_int(row.get('DIAS_LABORADOS', 0))
            if days_worked <= 0:
                errors.append("DIAS_LABORADOS must be greater than 0")
            
            amount = safe_float(row.get('CANTIDAD', 0))
            if amount <= 0:
                errors.append("CANTIDAD must be greater than 0")
            
            payroll_type = safe_str(row.get('TIPO_NOMINA', 'regular'))
            notes = safe_str(row.get('NOTAS', ''))
            
            # Parse FECHA_REGISTRO if available (otherwise use current time)
            fecha_registro = parse_datetime(row.get('FECHA_REGISTRO'))
            
            # If there are errors, skip this row
            if errors:
                skipped_rows.append({
                    "row": row_num,
                    "operador": operador,
                    "sucursal": sucursal,
                    "errors": errors
                })
                continue
            
            # Add valid record
            valid_records.append({
                "worker_id": worker_id,
                "branch_id": branch_id,
                "date": payroll_date,
                "days_worked": days_worked,
                "amount": amount,
                "payroll_type": payroll_type if payroll_type else "regular",
                "notes": notes if notes else None,
                "created_at": fecha_registro if fecha_registro else datetime.now(timezone.utc),
                "review_state": "pending"
            })
        
        # If no valid records, return early
        if not valid_records:
            return {
                "success": False,
                "message": "No valid records to import",
                "total_rows": len(df),
                "imported": 0,
                "skipped": len(skipped_rows),
                "skipped_rows": skipped_rows
            }
        
        # Import valid records in a transaction
        try:
            imported_count = 0
            for record_data in valid_records:
                payroll_record = Payroll(**record_data)
                db.add(payroll_record)
                imported_count += 1
            
            db.commit()
            
            return {
                "success": True,
                "message": f"Successfully imported {imported_count} payroll records",
                "total_rows": len(df),
                "imported": imported_count,
                "skipped": len(skipped_rows),
                "skipped_rows": skipped_rows
            }
            
        except Exception as e:
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Database error during import: {str(e)}"
            )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing Excel file: {str(e)}"
        )
